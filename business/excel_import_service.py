"""
Excel Product Import Service.
Parses, validates, previews, and bulk-imports products from .xlsx files using openpyxl.
Auto-creates missing Brand and Category records, initializes current_stock to 0.
"""

import os
import re
import uuid
import logging
from typing import Optional, List, Dict, Any, Tuple, Set
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from data.db import DatabaseManager

logger = logging.getLogger(__name__)


class ExcelImportService:
    """
    Handles parsing, validation, auto-taxonomy creation, and bulk ingestion of products from Excel.
    """

    EXPECTED_COLUMNS = [
        "Product Name",
        "Brand",
        "Model",
        "Category",
        "Minimum Stock",
        "Warranty Duration",
    ]

    def __init__(self, db: Optional[DatabaseManager] = None):
        self.db = db or DatabaseManager.get_instance()

    def parse_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Parse and validate an Excel file without committing changes.
        Returns a preview payload with valid rows, error rows, new brands, and new categories.
        """
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "total_rows": 0,
                "valid_rows": [],
                "error_rows": [],
                "new_categories": [],
                "new_brands": [],
            }

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            if sheet is None:
                return {
                    "success": False,
                    "error": "Excel workbook contains no active sheet.",
                    "total_rows": 0,
                    "valid_rows": [],
                    "error_rows": [],
                    "new_categories": [],
                    "new_brands": [],
                }
        except Exception as e:
            logger.error(f"Failed to read Excel file '{file_path}': {e}", exc_info=True)
            return {
                "success": False,
                "error": f"Failed to open Excel file: {e}",
                "total_rows": 0,
                "valid_rows": [],
                "error_rows": [],
                "new_categories": [],
                "new_brands": [],
            }

        # 1. Read existing categories & brands from DB for fast in-memory matching
        existing_cats = {
            r["name"].strip().lower(): r["id"]
            for r in self.db.execute_query("SELECT id, name FROM categories WHERE is_active = 1;")
        }
        existing_brands = {
            r["name"].strip().lower(): r["id"]
            for r in self.db.execute_query("SELECT id, name FROM brands WHERE is_active = 1;")
        }

        # 2. Map Headers
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = None
        header_row_idx = 0

        for idx, row in enumerate(rows_iter, start=1):
            if any(cell is not None and str(cell).strip() for cell in row):
                header_row = row
                header_row_idx = idx
                break

        if not header_row:
            return {
                "success": False,
                "error": "The Excel file is completely empty.",
                "total_rows": 0,
                "valid_rows": [],
                "error_rows": [],
                "new_categories": [],
                "new_brands": [],
            }

        col_map = self._map_headers(header_row)
        if "name" not in col_map:
            return {
                "success": False,
                "error": (
                    "Could not find required column 'Product Name' in header row. "
                    f"Found headers: {[str(c).strip() for c in header_row if c is not None]}. "
                    "Expected columns: Product Name, Brand, Model, Category, Minimum Stock, Warranty Duration."
                ),
                "total_rows": 0,
                "valid_rows": [],
                "error_rows": [],
                "new_categories": [],
                "new_brands": [],
            }

        valid_rows: List[Dict[str, Any]] = []
        error_rows: List[Dict[str, Any]] = []
        new_categories_set: Set[str] = set()
        new_brands_set: Set[str] = set()

        total_data_rows = 0

        # 3. Process Data Rows
        for excel_row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            # Check if row is completely blank
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            total_data_rows += 1
            row_errors: List[str] = []

            # --- Extract raw cell values ---
            raw_name = self._get_cell_value(row, col_map.get("name"))
            raw_brand = self._get_cell_value(row, col_map.get("brand"))
            raw_model = self._get_cell_value(row, col_map.get("model"))
            raw_category = self._get_cell_value(row, col_map.get("category"))
            raw_min_stock = self._get_cell_value(row, col_map.get("min_stock"))
            raw_warranty = self._get_cell_value(row, col_map.get("warranty"))

            # --- Validate Product Name ---
            product_name = str(raw_name).strip() if raw_name is not None else ""
            if not product_name:
                row_errors.append("Product Name is required and cannot be blank.")

            # --- Validate Brand ---
            brand_name = str(raw_brand).strip() if raw_brand is not None else None
            is_new_brand = False
            brand_id = None
            if brand_name:
                b_key = brand_name.lower()
                if b_key in existing_brands:
                    brand_id = existing_brands[b_key]
                else:
                    is_new_brand = True
                    new_brands_set.add(brand_name)

            # --- Validate Model ---
            model = str(raw_model).strip() if raw_model is not None else None

            # --- Validate Category ---
            category_name = str(raw_category).strip() if raw_category is not None else None
            is_new_category = False
            category_id = None
            if category_name:
                c_key = category_name.lower()
                if c_key in existing_cats:
                    category_id = existing_cats[c_key]
                else:
                    is_new_category = True
                    new_categories_set.add(category_name)

            # --- Validate Minimum Stock ---
            min_stock = 5  # default
            if raw_min_stock is not None and str(raw_min_stock).strip() != "":
                clean_ms = str(raw_min_stock).strip().replace(",", "")
                # Remove suffix like 'units' or 'pcs'
                clean_ms = re.sub(r"[^\d.-]", "", clean_ms)
                try:
                    val_ms = float(clean_ms)
                    if val_ms < 0:
                        row_errors.append(f"Minimum Stock must be a non-negative number (got '{raw_min_stock}').")
                    else:
                        min_stock = int(round(val_ms))
                except ValueError:
                    row_errors.append(f"Minimum Stock must be a valid number (got '{raw_min_stock}').")

            # --- Validate Warranty Duration ---
            warranty_months = 0  # default
            if raw_warranty is not None and str(raw_warranty).strip() != "":
                clean_w = str(raw_warranty).strip().replace(",", "")
                # Remove suffix like 'months' or 'mos' or 'yr'
                clean_w = re.sub(r"[^\d.-]", "", clean_w)
                try:
                    val_w = float(clean_w)
                    if val_w < 0:
                        row_errors.append(f"Warranty Duration must be a non-negative number of months (got '{raw_warranty}').")
                    else:
                        warranty_months = int(round(val_w))
                except ValueError:
                    row_errors.append(f"Warranty Duration must be a valid number of months (got '{raw_warranty}').")

            row_data = {
                "row_number": excel_row_idx,
                "name": product_name,
                "brand_name": brand_name,
                "brand_id": brand_id,
                "is_new_brand": is_new_brand,
                "model": model,
                "category_name": category_name,
                "category_id": category_id,
                "is_new_category": is_new_category,
                "min_stock_alert": min_stock,
                "warranty_period_months": warranty_months,
                "current_stock": 0,  # System-managed: always 0 upon import
            }

            if row_errors:
                row_data["errors"] = row_errors
                row_data["error_reason"] = " • ".join(row_errors)
                error_rows.append(row_data)
            else:
                valid_rows.append(row_data)

        return {
            "success": True,
            "file_name": os.path.basename(file_path),
            "file_path": file_path,
            "total_rows": total_data_rows,
            "valid_count": len(valid_rows),
            "error_count": len(error_rows),
            "valid_rows": valid_rows,
            "error_rows": error_rows,
            "new_categories": sorted(list(new_categories_set)),
            "new_brands": sorted(list(new_brands_set)),
        }

    def commit_import(self, valid_rows: List[Dict[str, Any]]) -> Tuple[bool, int, int, int, str]:
        """
        Execute atomic bulk import of valid product rows.
        Auto-creates missing categories and brands.
        Returns: (success, imported_count, new_categories_count, new_brands_count, message)
        """
        if not valid_rows:
            return False, 0, 0, 0, "No valid product rows to import."

        imported_products_count = 0
        new_cats_count = 0
        new_brands_count = 0

        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()

                # 1. Fetch current brand and category maps
                cursor.execute("SELECT id, name FROM brands WHERE is_active = 1;")
                brand_map = {r["name"].strip().lower(): r["id"] for r in cursor.fetchall()}

                cursor.execute("SELECT id, name FROM categories WHERE is_active = 1;")
                cat_map = {r["name"].strip().lower(): r["id"] for r in cursor.fetchall()}

                # 2. Auto-create missing Brands
                for row in valid_rows:
                    b_name = row.get("brand_name")
                    if b_name and b_name.strip():
                        b_key = b_name.strip().lower()
                        if b_key not in brand_map:
                            cursor.execute(
                                "INSERT INTO brands (name, is_active) VALUES (?, 1);",
                                (b_name.strip(),),
                            )
                            new_brand_id = cursor.lastrowid
                            brand_map[b_key] = new_brand_id
                            new_brands_count += 1
                            logger.info(f"Auto-created brand: '{b_name}' with ID {new_brand_id}")

                # 3. Auto-create missing Categories
                for row in valid_rows:
                    c_name = row.get("category_name")
                    if c_name and c_name.strip():
                        c_key = c_name.strip().lower()
                        if c_key not in cat_map:
                            cursor.execute(
                                "INSERT INTO categories (name, is_active) VALUES (?, 1);",
                                (c_name.strip(),),
                            )
                            new_cat_id = cursor.lastrowid
                            cat_map[c_key] = new_cat_id
                            new_cats_count += 1
                            logger.info(f"Auto-created category: '{c_name}' with ID {new_cat_id}")

                # 4. Bulk Insert Products
                for row in valid_rows:
                    name = row["name"]
                    model = row.get("model")
                    b_name = row.get("brand_name")
                    c_name = row.get("category_name")
                    min_stock = row.get("min_stock_alert", 5)
                    warranty = row.get("warranty_period_months", 0)

                    resolved_brand_id = brand_map.get(b_name.strip().lower()) if b_name else None
                    resolved_cat_id = cat_map.get(c_name.strip().lower()) if c_name else None

                    # Generate system internal SKU
                    internal_sku = f"PRD-IMP-{uuid.uuid4().hex[:8].upper()}"

                    cursor.execute("""
                        INSERT INTO products (
                            sku, name, model, brand_id, category_id,
                            current_stock, min_stock_alert, warranty_period_months,
                            cost_price, selling_price, is_active, is_deleted
                        ) VALUES (?, ?, ?, ?, ?, 0, ?, ?, 0.0, 0.0, 1, 0);
                    """, (
                        internal_sku,
                        name,
                        model,
                        resolved_brand_id,
                        resolved_cat_id,
                        min_stock,
                        warranty,
                    ))
                    imported_products_count += 1

                conn.commit()

            msg = (
                f"Successfully imported {imported_products_count} products. "
                f"Created {new_cats_count} new categories and {new_brands_count} new brands."
            )
            logger.info(msg)
            return True, imported_products_count, new_cats_count, new_brands_count, msg

        except Exception as e:
            logger.error(f"Bulk product import failed: {e}", exc_info=True)
            return False, 0, 0, 0, f"Database transaction error during import: {e}"

    def generate_template(self, output_path: str) -> bool:
        """
        Generate a styled Excel template file for product import.
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products Import"

            # Header Style
            header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            headers = [
                "Product Name",
                "Brand",
                "Model",
                "Category",
                "Minimum Stock",
                "Warranty Duration",
            ]

            ws.append(headers)

            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            ws.row_dimensions[1].height = 28

            # Sample Data Rows
            sample_rows = [
                ["Samsung Galaxy S24 Ultra", "Samsung", "SM-S928B 256GB Titanium", "Smartphones & Tablets", 5, 12],
                ["Apple MacBook Air 15-inch M3", "Apple", "A3114 Midnight 16GB", "Laptops & Desktops", 3, 12],
                ["Sony WH-1000XM5 Wireless Headset", "Sony", "WH-1000XM5 Black", "Audio & Headphones", 8, 12],
                ["Dell UltraSharp 32-inch 4K Monitor", "Dell", "U3223QE 4K IPS Black", "Monitors & Displays", 2, 36],
                ["Logitech MX Master 3S Wireless Mouse", "Logitech", "MX Master 3S Graphite", "Computer Accessories", 10, 24],
            ]

            data_font = Font(name="Segoe UI", size=10)
            for row_data in sample_rows:
                ws.append(row_data)
                row_idx = ws.max_row
                for col_idx in range(1, len(row_data) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = data_font
                    if col_idx in (5, 6):
                        c.alignment = Alignment(horizontal="right")

            # Column Widths
            col_widths = {1: 34, 2: 18, 3: 28, 4: 24, 5: 18, 6: 22}
            for col_num, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = width

            wb.save(output_path)
            logger.info(f"Generated product import template at: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to generate template '{output_path}': {e}", exc_info=True)
            return False

    def _map_headers(self, header_row: Tuple[Any, ...]) -> Dict[str, int]:
        """
        Map Excel header cell names to standard field keys.
        """
        mapping: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            text = str(cell).strip().lower()
            clean_text = re.sub(r"[_\-\s\(\)]+", " ", text).strip()

            if clean_text in ("product name", "product", "name", "item name", "item", "product title", "title"):
                mapping["name"] = idx
            elif clean_text in ("brand", "brand name", "manufacturer", "make", "brand manufacturer"):
                mapping["brand"] = idx
            elif clean_text in ("model", "model no", "model number", "model name", "model num"):
                mapping["model"] = idx
            elif clean_text in ("category", "category name", "product category", "type", "group", "category type"):
                mapping["category"] = idx
            elif clean_text in ("minimum stock", "min stock", "min stock alert", "minimum stock level", "min stock level", "low stock alert", "min stock qty", "minimum stock quantity", "min stock units", "minimum stock units"):
                mapping["min_stock"] = idx
            elif clean_text in ("warranty duration", "warranty", "warranty months", "warranty duration months", "warranty period", "warranty period months", "warranty duration in months"):
                mapping["warranty"] = idx

        return mapping

    def _get_cell_value(self, row: Tuple[Any, ...], col_idx: Optional[int]) -> Any:
        """Safely retrieve cell value by column index."""
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]
